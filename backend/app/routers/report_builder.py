import io
import uuid
from datetime import datetime
from html import escape
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import Column, String, DateTime, Text, Boolean, select, text
from sqlalchemy.dialects.postgresql import UUID, JSONB
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import AppBase, get_app_db, get_odoo_db
from app.auth.dependencies import get_current_user, require_permission
from app.auth.models import User


class SavedReport(AppBase):
    __tablename__ = "saved_reports"

    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    user_id = Column(UUID(as_uuid=True), nullable=False, index=True)
    name = Column(String(200), nullable=False)
    description = Column(Text, nullable=True)
    config = Column(JSONB, nullable=False)
    is_shared = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


router = APIRouter(prefix="/api/report-builder", tags=["Report Builder"])

# Join definitions: each join has a unique key, SQL clause, alias, and exposed columns
JOIN_DEFS = {
    "partner": {
        "label": "Customer / Vendor",
        "alias": "rp",
        "sql": "LEFT JOIN res_partner rp ON {table}.partner_id = rp.id",
        "columns": [
            {"key": "rp.name", "label": "Partner Name", "type": "string", "as": "partner_name"},
            {"key": "rp.email", "label": "Partner Email", "type": "string", "as": "partner_email"},
            {"key": "rp.phone", "label": "Partner Phone", "type": "string", "as": "partner_phone"},
            {"key": "rp.city", "label": "Partner City", "type": "string", "as": "partner_city"},
        ],
    },
    "product": {
        "label": "Product",
        "alias": "pt",
        "sql": (
            "LEFT JOIN product_product pp ON {table}.product_id = pp.id "
            "LEFT JOIN product_template pt ON pp.product_tmpl_id = pt.id"
        ),
        "columns": [
            {"key": "pt.name->>'en_US'", "label": "Product Name", "type": "string", "as": "product_name"},
            {"key": "pt.default_code", "label": "Internal Ref", "type": "string", "as": "product_ref"},
            {"key": "pt.list_price", "label": "List Price", "type": "number", "as": "product_list_price"},
        ],
    },
    "location": {
        "label": "Location",
        "alias": "sl",
        "sql": "LEFT JOIN stock_location sl ON {table}.location_id = sl.id",
        "columns": [
            {"key": "sl.complete_name", "label": "Location", "type": "string", "as": "location_name"},
        ],
    },
    "project": {
        "label": "Project",
        "alias": "pp",
        "sql": "LEFT JOIN project_project pp ON {table}.project_id = pp.id",
        "columns": [
            {"key": "pp.name->>'en_US'", "label": "Project Name", "type": "string", "as": "project_name"},
        ],
    },
    "order_lines": {
        "label": "Order Lines",
        "alias": "sol",
        "sql": "INNER JOIN sale_order_line sol ON sol.order_id = {table}.id",
        "columns": [
            {"key": "sol.product_uom_qty", "label": "Line Qty", "type": "number", "as": "line_qty"},
            {"key": "sol.price_unit", "label": "Line Unit Price", "type": "number", "as": "line_unit_price"},
            {"key": "sol.price_subtotal", "label": "Line Subtotal", "type": "number", "as": "line_subtotal"},
            {"key": "sol.margin", "label": "Line Margin", "type": "number", "as": "line_margin"},
        ],
    },
    "po_lines": {
        "label": "PO Lines",
        "alias": "pol",
        "sql": "INNER JOIN purchase_order_line pol ON pol.order_id = {table}.id",
        "columns": [
            {"key": "pol.product_qty", "label": "Line Qty", "type": "number", "as": "po_line_qty"},
            {"key": "pol.price_unit", "label": "Line Unit Price", "type": "number", "as": "po_line_unit_price"},
            {"key": "pol.price_subtotal", "label": "Line Subtotal", "type": "number", "as": "po_line_subtotal"},
        ],
    },
    "move_lines": {
        "label": "Journal Items",
        "alias": "aml",
        "sql": "INNER JOIN account_move_line aml ON aml.move_id = {table}.id",
        "columns": [
            {"key": "aml.debit", "label": "Debit", "type": "number", "as": "jrnl_debit"},
            {"key": "aml.credit", "label": "Credit", "type": "number", "as": "jrnl_credit"},
            {"key": "aml.balance", "label": "Balance", "type": "number", "as": "jrnl_balance"},
        ],
    },
}

AVAILABLE_TABLES = {
    "sale_order": {
        "label": "Sales Orders",
        "columns": [
            {"key": "name", "label": "Order Number", "type": "string"},
            {"key": "date_order", "label": "Order Date", "type": "datetime"},
            {"key": "state", "label": "Status", "type": "string"},
            {"key": "amount_untaxed", "label": "Untaxed Amount", "type": "number"},
            {"key": "amount_tax", "label": "Tax", "type": "number"},
            {"key": "amount_total", "label": "Total", "type": "number"},
            {"key": "invoice_status", "label": "Invoice Status", "type": "string"},
        ],
        "available_joins": ["partner", "order_lines"],
    },
    "purchase_order": {
        "label": "Purchase Orders",
        "columns": [
            {"key": "name", "label": "PO Number", "type": "string"},
            {"key": "date_order", "label": "Order Date", "type": "datetime"},
            {"key": "state", "label": "Status", "type": "string"},
            {"key": "amount_untaxed", "label": "Untaxed Amount", "type": "number"},
            {"key": "amount_tax", "label": "Tax", "type": "number"},
            {"key": "amount_total", "label": "Total", "type": "number"},
        ],
        "available_joins": ["partner", "po_lines"],
    },
    "account_move": {
        "label": "Journal Entries",
        "columns": [
            {"key": "name", "label": "Entry Number", "type": "string"},
            {"key": "date", "label": "Date", "type": "date"},
            {"key": "move_type", "label": "Type", "type": "string"},
            {"key": "state", "label": "State", "type": "string"},
            {"key": "amount_total", "label": "Total", "type": "number"},
            {"key": "amount_residual", "label": "Amount Due", "type": "number"},
            {"key": "payment_state", "label": "Payment State", "type": "string"},
        ],
        "available_joins": ["partner", "move_lines"],
    },
    "stock_quant": {
        "label": "Stock Quantities",
        "columns": [
            {"key": "quantity", "label": "Quantity", "type": "number"},
            {"key": "reserved_quantity", "label": "Reserved", "type": "number"},
        ],
        "available_joins": ["product", "location"],
    },
    "mrp_production": {
        "label": "Manufacturing Orders",
        "columns": [
            {"key": "name", "label": "MO Number", "type": "string"},
            {"key": "state", "label": "Status", "type": "string"},
            {"key": "priority", "label": "Priority", "type": "string"},
            {"key": "product_qty", "label": "Quantity", "type": "number"},
            {"key": "origin", "label": "Source", "type": "string"},
            {"key": "date_start", "label": "Start Date", "type": "datetime"},
            {"key": "date_finished", "label": "Finished Date", "type": "datetime"},
            {"key": "create_date", "label": "Created", "type": "datetime"},
        ],
        "available_joins": ["product"],
    },
    "helpdesk_ticket": {
        "label": "Helpdesk Tickets",
        "columns": [
            {"key": "ticket_ref", "label": "Ticket #", "type": "string"},
            {"key": "name", "label": "Subject", "type": "string"},
            {"key": "partner_name", "label": "Contact", "type": "string"},
            {"key": "partner_email", "label": "Email", "type": "string"},
            {"key": "priority", "label": "Priority", "type": "string"},
            {"key": "create_date", "label": "Created", "type": "datetime"},
            {"key": "close_date", "label": "Closed", "type": "datetime"},
        ],
        "available_joins": [],
    },
    "crm_lead": {
        "label": "CRM Leads",
        "columns": [
            {"key": "name", "label": "Lead Name", "type": "string"},
            {"key": "partner_name", "label": "Contact", "type": "string"},
            {"key": "email_from", "label": "Email", "type": "string"},
            {"key": "expected_revenue", "label": "Expected Revenue", "type": "number"},
            {"key": "probability", "label": "Probability", "type": "number"},
            {"key": "priority", "label": "Priority", "type": "string"},
            {"key": "create_date", "label": "Created", "type": "datetime"},
            {"key": "date_closed", "label": "Closed", "type": "datetime"},
        ],
        "available_joins": ["partner"],
    },
    "project_task": {
        "label": "Project Tasks",
        "columns": [
            {"key": "name", "label": "Task Name", "type": "string"},
            {"key": "state", "label": "State", "type": "string"},
            {"key": "priority", "label": "Priority", "type": "string"},
            {"key": "date_deadline", "label": "Deadline", "type": "datetime"},
            {"key": "allocated_hours", "label": "Allocated Hours", "type": "number"},
            {"key": "effective_hours", "label": "Effective Hours", "type": "number"},
            {"key": "progress", "label": "Progress %", "type": "number"},
            {"key": "create_date", "label": "Created", "type": "datetime"},
        ],
        "available_joins": ["project"],
    },
    "res_partner": {
        "label": "Customers",
        "columns": [
            {"key": "name", "label": "Name", "type": "string"},
            {"key": "email", "label": "Email", "type": "string"},
            {"key": "phone", "label": "Phone", "type": "string"},
            {"key": "city", "label": "City", "type": "string"},
            {"key": "customer_rank", "label": "Customer Rank", "type": "number"},
            {"key": "supplier_rank", "label": "Supplier Rank", "type": "number"},
            {"key": "create_date", "label": "Created", "type": "datetime"},
        ],
        "available_joins": [],
    },
}

AGGREGATIONS = ["count", "sum", "avg", "min", "max"]
CONDITIONS = ["eq", "neq", "gt", "gte", "lt", "lte", "contains", "between"]


class ReportFilter(BaseModel):
    column: str
    condition: str
    value: str
    value2: str | None = None


class ReportConfig(BaseModel):
    table: str
    columns: list[str]
    filters: list[ReportFilter] = []
    joins: list[str] = []  # join keys from JOIN_DEFS
    group_by: str | None = None
    aggregation: str | None = None
    agg_column: str | None = None
    order_by: str | None = None
    order_dir: str = "desc"
    limit: int = 100


class SavedReportCreate(BaseModel):
    name: str
    description: str | None = None
    config: ReportConfig
    is_shared: bool = False


class SavedReportResponse(BaseModel):
    id: str
    name: str
    description: str | None
    config: dict
    is_shared: bool
    created_at: datetime

    model_config = {"from_attributes": True}


@router.get("/tables")
async def list_tables():
    result = {}
    for k, v in AVAILABLE_TABLES.items():
        joins = []
        for jk in v.get("available_joins", []):
            jdef = JOIN_DEFS.get(jk)
            if jdef:
                joins.append({
                    "key": jk,
                    "label": jdef["label"],
                    "columns": [
                        {"key": c.get("as", c["key"]), "label": c["label"], "type": c["type"]}
                        for c in jdef["columns"]
                    ],
                })
        result[k] = {"label": v["label"], "columns": v["columns"], "joins": joins}
    return result


def _build_report_query(body: ReportConfig) -> tuple[str, dict, list[str]]:
    """Build SQL query from ReportConfig. Returns (sql, params, select_parts)."""
    if body.table not in AVAILABLE_TABLES:
        raise HTTPException(status_code=400, detail=f"Unknown table: {body.table}")

    table_def = AVAILABLE_TABLES[body.table]
    valid_cols = {c["key"] for c in table_def["columns"]}

    # Resolve active joins and build valid joined column set
    available_joins = set(table_def.get("available_joins", []))
    active_joins: list[str] = []
    joined_col_map: dict[str, str] = {}  # aliased_key -> sql_expr
    for jk in body.joins:
        if jk not in available_joins or jk not in JOIN_DEFS:
            continue
        active_joins.append(jk)
        jdef = JOIN_DEFS[jk]
        for c in jdef["columns"]:
            alias = c.get("as", c["key"])
            joined_col_map[alias] = c["key"]

    # Validate columns
    for col in body.columns:
        if col not in valid_cols and col not in joined_col_map:
            raise HTTPException(status_code=400, detail=f"Invalid column: {col}")

    select_parts = []
    if body.group_by and body.aggregation and body.agg_column:
        # Group-by column can be from base table or joined
        if body.group_by in valid_cols:
            select_parts.append(f"{body.table}.{body.group_by}")
        elif body.group_by in joined_col_map:
            expr = joined_col_map[body.group_by]
            select_parts.append(f"{expr} as {body.group_by}")
        else:
            select_parts.append(f"{body.table}.{body.group_by}")
        # Agg column
        if body.agg_column in valid_cols:
            select_parts.append(f"{body.aggregation}({body.table}.{body.agg_column}) as agg_value")
        elif body.agg_column in joined_col_map:
            expr = joined_col_map[body.agg_column]
            select_parts.append(f"{body.aggregation}({expr}) as agg_value")
        else:
            select_parts.append(f"{body.aggregation}({body.table}.{body.agg_column}) as agg_value")
    else:
        for col in body.columns:
            if col in joined_col_map:
                expr = joined_col_map[col]
                select_parts.append(f"{expr} as {col}")
            else:
                select_parts.append(f"{body.table}.{col}")

    # Build FROM + JOINs
    from_clause = body.table
    for jk in active_joins:
        jdef = JOIN_DEFS[jk]
        from_clause += " " + jdef["sql"].format(table=body.table)

    where_parts = []
    params = {}
    all_valid = valid_cols | set(joined_col_map.keys())
    for i, f in enumerate(body.filters):
        if f.column not in all_valid:
            continue
        param_key = f"p{i}"
        if f.column in joined_col_map:
            col_ref = joined_col_map[f.column]
        else:
            col_ref = f"{body.table}.{f.column}"
        if f.condition == "eq":
            where_parts.append(f"{col_ref} = :{param_key}")
            params[param_key] = f.value
        elif f.condition == "neq":
            where_parts.append(f"{col_ref} != :{param_key}")
            params[param_key] = f.value
        elif f.condition == "gt":
            where_parts.append(f"{col_ref} > :{param_key}")
            params[param_key] = float(f.value)
        elif f.condition == "gte":
            where_parts.append(f"{col_ref} >= :{param_key}")
            params[param_key] = float(f.value)
        elif f.condition == "lt":
            where_parts.append(f"{col_ref} < :{param_key}")
            params[param_key] = float(f.value)
        elif f.condition == "lte":
            where_parts.append(f"{col_ref} <= :{param_key}")
            params[param_key] = float(f.value)
        elif f.condition == "contains":
            where_parts.append(f"{col_ref}::text ILIKE :{param_key}")
            params[param_key] = f"%{f.value}%"
        elif f.condition == "between" and f.value2:
            where_parts.append(f"{col_ref} BETWEEN :{param_key}_a AND :{param_key}_b")
            params[f"{param_key}_a"] = f.value
            params[f"{param_key}_b"] = f.value2

    sql = f"SELECT {', '.join(select_parts)} FROM {from_clause}"
    if where_parts:
        sql += f" WHERE {' AND '.join(where_parts)}"
    if body.group_by and body.aggregation:
        if body.group_by in joined_col_map:
            sql += f" GROUP BY {joined_col_map[body.group_by]}"
        else:
            sql += f" GROUP BY {body.table}.{body.group_by}"
    if body.order_by and body.order_by in all_valid:
        if body.order_by in joined_col_map:
            order_expr = joined_col_map[body.order_by]
        else:
            order_expr = f"{body.table}.{body.order_by}"
        sql += f" ORDER BY {order_expr} {body.order_dir}"
    sql += f" LIMIT {min(body.limit, 5000)}"

    return sql, params, select_parts


@router.post("/query")
async def run_query(
    body: ReportConfig,
    odoo_db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("reports.builder")),
):
    sql, params, select_parts = _build_report_query(body)
    result = await odoo_db.execute(text(sql), params)
    rows = result.mappings().all()
    return {"columns": select_parts, "data": [dict(r) for r in rows], "total": len(rows)}


# --- Report Builder Export ---

BRAND_COLOR = "48CAE1"
_HEADER_FILL = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_CELL_FONT = Font(name="Arial", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E8EDF2"),
    right=Side(style="thin", color="E8EDF2"),
    top=Side(style="thin", color="E8EDF2"),
    bottom=Side(style="thin", color="E8EDF2"),
)


@router.post("/export/{fmt}")
async def export_report(
    fmt: Literal["csv", "excel", "pdf"],
    body: ReportConfig,
    odoo_db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("reports.builder")),
):
    sql, params, select_parts = _build_report_query(body)
    result = await odoo_db.execute(text(sql), params)
    rows = [dict(r) for r in result.mappings().all()]

    # Derive clean header labels from select parts
    headers = []
    keys = []
    for sp in select_parts:
        if " as " in sp:
            key = sp.split(" as ")[-1].strip()
        elif "." in sp:
            key = sp.split(".")[-1].strip()
        else:
            key = sp.strip()
        keys.append(key)
        headers.append(key.replace("_", " ").title())

    table_label = AVAILABLE_TABLES.get(body.table, {}).get("label", body.table)
    filename_base = f"report_{body.table}"

    if fmt == "csv":
        if not rows:
            return StreamingResponse(iter(["No data"]), media_type="text/csv")
        lines = [",".join(headers)]
        for row in rows:
            lines.append(",".join(str(row.get(k, "")).replace(",", ";") for k in keys))
        return StreamingResponse(
            iter(["\n".join(lines)]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    elif fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = table_label[:31]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(keys, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key)).font = _CELL_FONT
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    else:  # pdf
        from weasyprint import HTML as WeasyHTML
        header_cells = "".join(f"<th>{escape(h)}</th>" for h in headers)
        body_rows = ""
        for row in rows:
            cells = "".join(f"<td>{escape(str(row.get(k, '') or ''))}</td>" for k in keys)
            body_rows += f"<tr>{cells}</tr>\n"
        html_content = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{ size: A4 landscape; margin: 1.5cm; @bottom-right {{ content: "Page " counter(page) " of " counter(pages); font-size: 9px; color: #888; }} }}
body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #1a1a2e; }}
.header {{ display: flex; justify-content: space-between; margin-bottom: 16px; border-bottom: 3px solid #48CAE1; padding-bottom: 8px; }}
.title {{ font-size: 18px; font-weight: bold; color: #1a1a2e; }}
.count {{ font-size: 10px; color: #888; text-align: right; margin-bottom: 4px; }}
table {{ width: 100%; border-collapse: collapse; }}
th {{ background: #48CAE1; color: #fff; font-weight: bold; text-align: left; padding: 6px 8px; font-size: 10px; }}
td {{ padding: 5px 8px; border-bottom: 1px solid #e8edf2; font-size: 10px; }}
tr:nth-child(even) td {{ background: #f8fafc; }}
</style></head><body>
<div class="header"><div class="title">{escape(table_label)} Report</div></div>
<div class="count">{len(rows):,} record{"s" if len(rows) != 1 else ""}</div>
<table><thead><tr>{header_cells}</tr></thead><tbody>{body_rows}</tbody></table>
</body></html>"""
        pdf_bytes = WeasyHTML(string=html_content).write_pdf()
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
        )


@router.get("/saved", response_model=list[SavedReportResponse])
async def list_saved_reports(
    db: AsyncSession = Depends(get_app_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(SavedReport).where(
            (SavedReport.user_id == user.id) | (SavedReport.is_shared == True)
        ).order_by(SavedReport.updated_at.desc())
    )
    return result.scalars().all()


@router.post("/saved", response_model=SavedReportResponse, status_code=201)
async def save_report(
    body: SavedReportCreate,
    db: AsyncSession = Depends(get_app_db),
    user: User = Depends(get_current_user),
):
    report = SavedReport(
        user_id=user.id,
        name=body.name,
        description=body.description,
        config=body.config.model_dump(),
        is_shared=body.is_shared,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)
    return report


@router.delete("/saved/{report_id}", status_code=204)
async def delete_saved_report(
    report_id: str,
    db: AsyncSession = Depends(get_app_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(SavedReport).where(SavedReport.id == uuid.UUID(report_id))
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if report.user_id != user.id and not user.is_superadmin:
        raise HTTPException(status_code=403, detail="Not allowed")
    await db.delete(report)
    await db.commit()
