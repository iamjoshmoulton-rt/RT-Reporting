import io
from datetime import date
from html import escape

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from weasyprint import HTML

from app.database import get_odoo_db
from app.auth.dependencies import require_permission
from app.auth.models import User
from app.services import (
    sales_service, procurement_service, accounting_service, inventory_service,
    manufacturing_service, helpdesk_service, crm_service, projects_service, customers_service,
)

router = APIRouter(prefix="/api/export", tags=["Export"])

BRAND_COLOR = "48CAE1"
HEADER_FILL = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="E8EDF2"),
    right=Side(style="thin", color="E8EDF2"),
    top=Side(style="thin", color="E8EDF2"),
    bottom=Side(style="thin", color="E8EDF2"),
)


def _style_worksheet(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)


def _workbook_to_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(rows: list[dict], filename: str) -> StreamingResponse:
    if not rows:
        return StreamingResponse(iter(["No data"]), media_type="text/csv")

    headers = list(rows[0].keys())
    lines = [",".join(headers)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")).replace(",", ";") for h in headers))

    content = "\n".join(lines)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(
    rows: list[dict],
    headers: list[str],
    keys: list[str],
    title: str,
    filename: str,
    date_from: date | None = None,
    date_to: date | None = None,
) -> StreamingResponse:
    """Generate a styled PDF table report using WeasyPrint."""
    date_range = ""
    if date_from and date_to:
        date_range = f"{date_from.isoformat()} to {date_to.isoformat()}"
    elif date_from:
        date_range = f"From {date_from.isoformat()}"
    elif date_to:
        date_range = f"Through {date_to.isoformat()}"

    header_cells = "".join(f"<th>{escape(h)}</th>" for h in headers)
    body_rows = ""
    for row in rows:
        cells = "".join(
            f"<td>{escape(str(row.get(k, '') or ''))}</td>" for k in keys
        )
        body_rows += f"<tr>{cells}</tr>\n"

    html_content = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{ size: A4 landscape; margin: 1.5cm; @bottom-right {{ content: "Page " counter(page) " of " counter(pages); font-size: 9px; color: #888; }} }}
body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #1a1a2e; }}
.header {{ display: flex; justify-content: space-between; align-items: flex-end; margin-bottom: 16px; border-bottom: 3px solid #48CAE1; padding-bottom: 8px; }}
.title {{ font-size: 18px; font-weight: bold; color: #1a1a2e; }}
.subtitle {{ font-size: 11px; color: #666; margin-top: 2px; }}
.count {{ font-size: 10px; color: #888; text-align: right; margin-bottom: 4px; }}
table {{ width: 100%; border-collapse: collapse; }}
th {{ background: #48CAE1; color: #fff; font-weight: bold; text-align: left; padding: 6px 8px; font-size: 10px; }}
td {{ padding: 5px 8px; border-bottom: 1px solid #e8edf2; font-size: 10px; }}
tr:nth-child(even) td {{ background: #f8fafc; }}
</style></head><body>
<div class="header">
  <div><div class="title">{escape(title)}</div>
  {f'<div class="subtitle">{escape(date_range)}</div>' if date_range else ''}</div>
</div>
<div class="count">{len(rows):,} record{"s" if len(rows) != 1 else ""}</div>
<table><thead><tr>{header_cells}</tr></thead><tbody>{body_rows}</tbody></table>
</body></html>"""

    pdf_bytes = HTML(string=html_content).write_pdf()
    buffer = io.BytesIO(pdf_bytes)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Sales Exports ---

@router.get("/sales/csv")
async def export_sales_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("sales.export_csv")),
):
    result = await sales_service.get_sales_orders(db, date_from, date_to, limit=5000)
    return _csv_response(result["orders"], "sales_orders.csv")


@router.get("/sales/excel")
async def export_sales_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("sales.export_excel")),
):
    result = await sales_service.get_sales_orders(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Orders"

    headers = ["Order #", "Customer", "Date", "Untaxed", "Total", "Status", "Invoice Status"]
    _style_worksheet(ws, headers)

    for i, order in enumerate(result["orders"], 2):
        ws.cell(row=i, column=1, value=order.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=order.get("customer_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=order.get("date_order")).font = CELL_FONT
        ws.cell(row=i, column=4, value=order.get("amount_untaxed")).font = CELL_FONT
        ws.cell(row=i, column=5, value=order.get("amount_total")).font = CELL_FONT
        ws.cell(row=i, column=6, value=order.get("state")).font = CELL_FONT
        ws.cell(row=i, column=7, value=order.get("invoice_status")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "sales_orders.xlsx")


@router.get("/sales/kpi-drilldown")
async def export_sales_kpi_drilldown(
    kpi: str = Query(..., description="One of: invoiced_revenue, invoiced_margin, margin_percent, units_sold, open_pipeline, open_pipeline_date, max_potential_revenue, avg_sell_price"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("sales.export_excel")),
):
    if kpi not in sales_service.DRILLDOWN_KPI_TYPES:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Unknown KPI: {kpi}")
    result = await sales_service.get_sales_kpi_drilldown(db, kpi, date_from, date_to, offset=0, limit=10000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Drilldown"

    if result["type"] == "orders":
        headers = ["Order #", "Customer", "Order Date", "Commitment Date", "Untaxed", "Total", "Margin", "Status", "Invoice Status"]
        _style_worksheet(ws, headers)
        for i, row in enumerate(result["rows"], 2):
            ws.cell(row=i, column=1, value=row.get("name")).font = CELL_FONT
            ws.cell(row=i, column=2, value=row.get("customer_name")).font = CELL_FONT
            ws.cell(row=i, column=3, value=row.get("date_order")).font = CELL_FONT
            ws.cell(row=i, column=4, value=row.get("commitment_date")).font = CELL_FONT
            ws.cell(row=i, column=5, value=row.get("amount_untaxed")).font = CELL_FONT
            ws.cell(row=i, column=6, value=row.get("amount_total")).font = CELL_FONT
            ws.cell(row=i, column=7, value=row.get("margin")).font = CELL_FONT
            ws.cell(row=i, column=8, value=row.get("state")).font = CELL_FONT
            ws.cell(row=i, column=9, value=row.get("invoice_status")).font = CELL_FONT
    else:
        headers = ["Order #", "Product", "Customer", "Qty", "Unit Price", "Subtotal", "Margin", "Date"]
        _style_worksheet(ws, headers)
        for i, row in enumerate(result["rows"], 2):
            ws.cell(row=i, column=1, value=row.get("order_name")).font = CELL_FONT
            ws.cell(row=i, column=2, value=row.get("product_name")).font = CELL_FONT
            ws.cell(row=i, column=3, value=row.get("customer_name")).font = CELL_FONT
            ws.cell(row=i, column=4, value=row.get("qty")).font = CELL_FONT
            ws.cell(row=i, column=5, value=row.get("price_unit")).font = CELL_FONT
            ws.cell(row=i, column=6, value=row.get("subtotal")).font = CELL_FONT
            ws.cell(row=i, column=7, value=row.get("margin")).font = CELL_FONT
            ws.cell(row=i, column=8, value=row.get("date_order")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, f"{kpi}_drilldown.xlsx")


# --- Procurement Exports ---

@router.get("/procurement/csv")
async def export_procurement_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("procurement.export_csv")),
):
    result = await procurement_service.get_purchase_orders(db, date_from, date_to, limit=5000)
    return _csv_response(result["orders"], "purchase_orders.csv")


@router.get("/procurement/excel")
async def export_procurement_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("procurement.export_excel")),
):
    result = await procurement_service.get_purchase_orders(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = ["PO #", "Vendor", "Date", "Untaxed", "Total", "Status", "Invoice Status"]
    _style_worksheet(ws, headers)

    for i, order in enumerate(result["orders"], 2):
        ws.cell(row=i, column=1, value=order.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=order.get("vendor_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=order.get("date_order")).font = CELL_FONT
        ws.cell(row=i, column=4, value=order.get("amount_untaxed")).font = CELL_FONT
        ws.cell(row=i, column=5, value=order.get("amount_total")).font = CELL_FONT
        ws.cell(row=i, column=6, value=order.get("state")).font = CELL_FONT
        ws.cell(row=i, column=7, value=order.get("invoice_status")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "purchase_orders.xlsx")


# --- Accounting Exports ---

@router.get("/accounting/csv")
async def export_accounting_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    move_type: str | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("accounting.export_csv")),
):
    result = await accounting_service.get_invoices(db, date_from, date_to, move_type, limit=5000)
    return _csv_response(result["invoices"], "invoices.csv")


@router.get("/accounting/excel")
async def export_accounting_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    move_type: str | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("accounting.export_excel")),
):
    result = await accounting_service.get_invoices(db, date_from, date_to, move_type, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = ["Number", "Partner", "Type", "Date", "Due Date", "Total", "Amount Due", "Payment"]
    _style_worksheet(ws, headers)

    for i, inv in enumerate(result["invoices"], 2):
        ws.cell(row=i, column=1, value=inv.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=inv.get("partner_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=inv.get("move_type")).font = CELL_FONT
        ws.cell(row=i, column=4, value=inv.get("date")).font = CELL_FONT
        ws.cell(row=i, column=5, value=inv.get("invoice_date_due")).font = CELL_FONT
        ws.cell(row=i, column=6, value=inv.get("amount_total")).font = CELL_FONT
        ws.cell(row=i, column=7, value=inv.get("amount_residual")).font = CELL_FONT
        ws.cell(row=i, column=8, value=inv.get("payment_state")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "invoices.xlsx")


# --- Inventory Exports ---

@router.get("/inventory/csv")
async def export_inventory_csv(
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("inventory.export_csv")),
):
    result = await inventory_service.get_stock_levels(db, limit=5000)
    return _csv_response(result["items"], "stock_levels.csv")


@router.get("/inventory/excel")
async def export_inventory_excel(
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("inventory.export_excel")),
):
    result = await inventory_service.get_stock_levels(db, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Levels"

    headers = ["Ref", "Product", "On Hand", "Reserved", "Available"]
    _style_worksheet(ws, headers)

    for i, item in enumerate(result["items"], 2):
        ws.cell(row=i, column=1, value=item.get("internal_ref")).font = CELL_FONT
        ws.cell(row=i, column=2, value=item.get("product_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=item.get("on_hand")).font = CELL_FONT
        ws.cell(row=i, column=4, value=item.get("reserved")).font = CELL_FONT
        ws.cell(row=i, column=5, value=item.get("available")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "stock_levels.xlsx")


# --- Manufacturing Exports ---

@router.get("/manufacturing/csv")
async def export_manufacturing_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("manufacturing.export_csv")),
):
    result = await manufacturing_service.get_manufacturing_orders(db, date_from, date_to, limit=5000)
    return _csv_response(result["orders"], "manufacturing_orders.csv")


@router.get("/manufacturing/excel")
async def export_manufacturing_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("manufacturing.export_excel")),
):
    result = await manufacturing_service.get_manufacturing_orders(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Manufacturing Orders"

    headers = ["MO #", "Product", "Qty", "Status", "Source", "Start Date", "Finished", "Created"]
    _style_worksheet(ws, headers)

    for i, order in enumerate(result["orders"], 2):
        ws.cell(row=i, column=1, value=order.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=order.get("product_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=order.get("product_qty")).font = CELL_FONT
        ws.cell(row=i, column=4, value=order.get("state")).font = CELL_FONT
        ws.cell(row=i, column=5, value=order.get("origin")).font = CELL_FONT
        ws.cell(row=i, column=6, value=order.get("date_start")).font = CELL_FONT
        ws.cell(row=i, column=7, value=order.get("date_finished")).font = CELL_FONT
        ws.cell(row=i, column=8, value=order.get("create_date")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "manufacturing_orders.xlsx")


# --- Helpdesk Exports ---

@router.get("/helpdesk/csv")
async def export_helpdesk_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("helpdesk.export_csv")),
):
    result = await helpdesk_service.get_helpdesk_tickets(db, date_from, date_to, limit=5000)
    return _csv_response(result["tickets"], "helpdesk_tickets.csv")


@router.get("/helpdesk/excel")
async def export_helpdesk_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("helpdesk.export_excel")),
):
    result = await helpdesk_service.get_helpdesk_tickets(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Helpdesk Tickets"

    headers = ["Ticket #", "Subject", "Contact", "Email", "Team", "Stage", "Priority", "Created", "Closed"]
    _style_worksheet(ws, headers)

    for i, ticket in enumerate(result["tickets"], 2):
        ws.cell(row=i, column=1, value=ticket.get("ticket_ref")).font = CELL_FONT
        ws.cell(row=i, column=2, value=ticket.get("name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=ticket.get("partner_name")).font = CELL_FONT
        ws.cell(row=i, column=4, value=ticket.get("partner_email")).font = CELL_FONT
        ws.cell(row=i, column=5, value=ticket.get("team_name")).font = CELL_FONT
        ws.cell(row=i, column=6, value=ticket.get("stage_name")).font = CELL_FONT
        ws.cell(row=i, column=7, value=ticket.get("priority")).font = CELL_FONT
        ws.cell(row=i, column=8, value=ticket.get("create_date")).font = CELL_FONT
        ws.cell(row=i, column=9, value=ticket.get("close_date")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "helpdesk_tickets.xlsx")


# --- CRM Exports ---

@router.get("/crm/csv")
async def export_crm_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("crm.export_csv")),
):
    result = await crm_service.get_crm_leads(db, date_from, date_to, limit=5000)
    return _csv_response(result["leads"], "crm_leads.csv")


@router.get("/crm/excel")
async def export_crm_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("crm.export_excel")),
):
    result = await crm_service.get_crm_leads(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "CRM Leads"

    headers = ["Lead", "Contact", "Email", "Stage", "Expected Revenue", "Probability", "Priority", "Created", "Closed"]
    _style_worksheet(ws, headers)

    for i, lead in enumerate(result["leads"], 2):
        ws.cell(row=i, column=1, value=lead.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=lead.get("partner_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=lead.get("email_from")).font = CELL_FONT
        ws.cell(row=i, column=4, value=lead.get("stage_name")).font = CELL_FONT
        ws.cell(row=i, column=5, value=lead.get("expected_revenue")).font = CELL_FONT
        ws.cell(row=i, column=6, value=lead.get("probability")).font = CELL_FONT
        ws.cell(row=i, column=7, value=lead.get("priority")).font = CELL_FONT
        ws.cell(row=i, column=8, value=lead.get("create_date")).font = CELL_FONT
        ws.cell(row=i, column=9, value=lead.get("date_closed")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "crm_leads.xlsx")


# --- Projects Exports ---

@router.get("/projects/csv")
async def export_projects_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("projects.export_csv")),
):
    result = await projects_service.get_project_tasks(db, date_from, date_to, limit=5000)
    return _csv_response(result["tasks"], "project_tasks.csv")


@router.get("/projects/excel")
async def export_projects_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("projects.export_excel")),
):
    result = await projects_service.get_project_tasks(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Tasks"

    headers = ["Task", "Project", "Stage", "State", "Priority", "Deadline", "Allocated Hrs", "Effective Hrs", "Progress %", "Created"]
    _style_worksheet(ws, headers)

    for i, task in enumerate(result["tasks"], 2):
        ws.cell(row=i, column=1, value=task.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=task.get("project_name")).font = CELL_FONT
        ws.cell(row=i, column=3, value=task.get("stage_name")).font = CELL_FONT
        ws.cell(row=i, column=4, value=task.get("state")).font = CELL_FONT
        ws.cell(row=i, column=5, value=task.get("priority")).font = CELL_FONT
        ws.cell(row=i, column=6, value=task.get("date_deadline")).font = CELL_FONT
        ws.cell(row=i, column=7, value=task.get("allocated_hours")).font = CELL_FONT
        ws.cell(row=i, column=8, value=task.get("effective_hours")).font = CELL_FONT
        ws.cell(row=i, column=9, value=task.get("progress")).font = CELL_FONT
        ws.cell(row=i, column=10, value=task.get("create_date")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "project_tasks.xlsx")


# --- Customers Exports ---

@router.get("/customers/csv")
async def export_customers_csv(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("customers.export_csv")),
):
    result = await customers_service.get_customer_list(db, date_from, date_to, limit=5000)
    return _csv_response(result["customers"], "customers.csv")


@router.get("/customers/excel")
async def export_customers_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("customers.export_excel")),
):
    result = await customers_service.get_customer_list(db, date_from, date_to, limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["Customer", "Email", "Phone", "City", "Orders", "Total Spend", "Since"]
    _style_worksheet(ws, headers)

    for i, cust in enumerate(result["customers"], 2):
        ws.cell(row=i, column=1, value=cust.get("name")).font = CELL_FONT
        ws.cell(row=i, column=2, value=cust.get("email")).font = CELL_FONT
        ws.cell(row=i, column=3, value=cust.get("phone")).font = CELL_FONT
        ws.cell(row=i, column=4, value=cust.get("city")).font = CELL_FONT
        ws.cell(row=i, column=5, value=cust.get("order_count")).font = CELL_FONT
        ws.cell(row=i, column=6, value=cust.get("total_spend")).font = CELL_FONT
        ws.cell(row=i, column=7, value=cust.get("create_date")).font = CELL_FONT

    _auto_width(ws)
    return _workbook_to_response(wb, "customers.xlsx")


# =============================================
# PDF Exports (all modules)
# =============================================

@router.get("/sales/pdf")
async def export_sales_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("sales.export_pdf")),
):
    result = await sales_service.get_sales_orders(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["orders"],
        ["Order #", "Customer", "Date", "Untaxed", "Total", "Status", "Invoice"],
        ["name", "customer_name", "date_order", "amount_untaxed", "amount_total", "state", "invoice_status"],
        "Sales Orders", "sales_orders.pdf", date_from, date_to,
    )


@router.get("/procurement/pdf")
async def export_procurement_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("procurement.export_pdf")),
):
    result = await procurement_service.get_purchase_orders(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["orders"],
        ["PO #", "Vendor", "Date", "Untaxed", "Total", "Status", "Invoice"],
        ["name", "vendor_name", "date_order", "amount_untaxed", "amount_total", "state", "invoice_status"],
        "Purchase Orders", "purchase_orders.pdf", date_from, date_to,
    )


@router.get("/accounting/pdf")
async def export_accounting_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    move_type: str | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("accounting.export_pdf")),
):
    result = await accounting_service.get_invoices(db, date_from, date_to, move_type, limit=5000)
    return _pdf_response(
        result["invoices"],
        ["Number", "Partner", "Type", "Date", "Due Date", "Total", "Amount Due", "Payment"],
        ["name", "partner_name", "move_type", "date", "invoice_date_due", "amount_total", "amount_residual", "payment_state"],
        "Invoices & Bills", "invoices.pdf", date_from, date_to,
    )


@router.get("/inventory/pdf")
async def export_inventory_pdf(
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("inventory.export_pdf")),
):
    result = await inventory_service.get_stock_levels(db, limit=5000)
    return _pdf_response(
        result["items"],
        ["Ref", "Product", "On Hand", "Reserved", "Available"],
        ["internal_ref", "product_name", "on_hand", "reserved", "available"],
        "Stock Levels", "stock_levels.pdf",
    )


@router.get("/manufacturing/pdf")
async def export_manufacturing_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("manufacturing.export_pdf")),
):
    result = await manufacturing_service.get_manufacturing_orders(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["orders"],
        ["MO #", "Product", "Qty", "Status", "Source", "Start Date", "Finished"],
        ["name", "product_name", "product_qty", "state", "origin", "date_start", "date_finished"],
        "Manufacturing Orders", "manufacturing_orders.pdf", date_from, date_to,
    )


@router.get("/helpdesk/pdf")
async def export_helpdesk_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("helpdesk.export_pdf")),
):
    result = await helpdesk_service.get_helpdesk_tickets(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["tickets"],
        ["Ticket #", "Subject", "Contact", "Team", "Stage", "Priority", "Created", "Closed"],
        ["ticket_ref", "name", "partner_name", "team_name", "stage_name", "priority", "create_date", "close_date"],
        "Helpdesk Tickets", "helpdesk_tickets.pdf", date_from, date_to,
    )


@router.get("/crm/pdf")
async def export_crm_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("crm.export_pdf")),
):
    result = await crm_service.get_crm_leads(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["leads"],
        ["Lead", "Contact", "Email", "Stage", "Expected Revenue", "Probability", "Priority", "Created"],
        ["name", "partner_name", "email_from", "stage_name", "expected_revenue", "probability", "priority", "create_date"],
        "CRM Leads", "crm_leads.pdf", date_from, date_to,
    )


@router.get("/projects/pdf")
async def export_projects_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("projects.export_pdf")),
):
    result = await projects_service.get_project_tasks(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["tasks"],
        ["Task", "Project", "Stage", "State", "Priority", "Deadline", "Alloc Hrs", "Progress %"],
        ["name", "project_name", "stage_name", "state", "priority", "date_deadline", "allocated_hours", "progress"],
        "Project Tasks", "project_tasks.pdf", date_from, date_to,
    )


@router.get("/customers/pdf")
async def export_customers_pdf(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_odoo_db),
    _user: User = Depends(require_permission("customers.export_pdf")),
):
    result = await customers_service.get_customer_list(db, date_from, date_to, limit=5000)
    return _pdf_response(
        result["customers"],
        ["Customer", "Email", "Phone", "City", "Orders", "Total Spend", "Since"],
        ["name", "email", "phone", "city", "order_count", "total_spend", "create_date"],
        "Customers", "customers.pdf", date_from, date_to,
    )
