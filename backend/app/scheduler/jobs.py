import io
import logging
import uuid
from datetime import date, timedelta

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from sqlalchemy import select, text

from app.config import get_settings
from app.database import OdooSessionLocal, AppSessionLocal
from app.email.sender import send_email
from app.services import (
    sales_service, procurement_service, accounting_service, inventory_service,
    helpdesk_service, crm_service, manufacturing_service, projects_service,
)

logger = logging.getLogger(__name__)

TEMPLATE_DIR = "app/email/templates"
jinja_env = Environment(
    loader=FileSystemLoader(TEMPLATE_DIR),
    autoescape=select_autoescape(["html"]),
)


async def run_scheduled_report(report_id: str, report_type: str, recipients: list[str],
                                filters: dict, attachment_format: str):
    settings = get_settings()
    today = date.today()
    date_from = filters.get("date_from", str(today - timedelta(days=30)))
    date_to = filters.get("date_to", str(today))

    # Handle Report Builder saved queries separately
    if report_type == "report_builder":
        await _run_saved_report_builder(
            report_id, filters, recipients, attachment_format, date_from, date_to
        )
        return

    async with OdooSessionLocal() as odoo_db:
        context = {
            "report_title": f"{'Summary' if report_type == 'summary' else report_type.replace('_', ' ').title()} Report",
            "frontend_url": settings.frontend_url,
            "date_from": date_from,
            "date_to": date_to,
        }

        if report_type == "summary":
            context["sales"] = await sales_service.get_sales_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["procurement"] = await procurement_service.get_procurement_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["accounting"] = await accounting_service.get_accounting_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["inventory"] = await inventory_service.get_inventory_summary(odoo_db)
            template = jinja_env.get_template("summary_report.html")

        elif report_type == "sales":
            context["summary"] = await sales_service.get_sales_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["top_customers"] = await sales_service.get_sales_by_customer(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to), 10)
            orders = await sales_service.get_sales_orders(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to), limit=15)
            context["recent_orders"] = orders["orders"]
            template = jinja_env.get_template("sales_report.html")

        else:
            context["sales"] = await sales_service.get_sales_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["procurement"] = await procurement_service.get_procurement_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["accounting"] = await accounting_service.get_accounting_summary(odoo_db, date.fromisoformat(date_from), date.fromisoformat(date_to))
            context["inventory"] = await inventory_service.get_inventory_summary(odoo_db)
            template = jinja_env.get_template("summary_report.html")

        html_body = template.render(**context)

        attachments = []
        if attachment_format in ("excel", "both"):
            excel_bytes = await _generate_excel_attachment(odoo_db, report_type, date_from, date_to)
            if excel_bytes:
                attachments.append((f"{report_type}_report.xlsx", excel_bytes,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

    subject = f"RT Reporting - {report_type.replace('_', ' ').title()} Report ({date_from} to {date_to})"

    try:
        await send_email(recipients, subject, html_body, attachments or None)
        logger.info(f"Sent {report_type} report to {recipients}")
    except Exception as e:
        logger.error(f"Failed to send report: {e}")


async def _run_saved_report_builder(
    report_id: str, filters: dict, recipients: list[str],
    attachment_format: str, date_from: str, date_to: str,
):
    """Run a saved Report Builder query and email the results as an Excel attachment."""
    from app.routers.report_builder import SavedReport, ReportConfig, _build_report_query, AVAILABLE_TABLES

    saved_report_id = filters.get("saved_report_id")
    if not saved_report_id:
        logger.error(f"Scheduled report {report_id}: no saved_report_id in filters")
        return

    # Load the saved report config
    async with AppSessionLocal() as app_db:
        result = await app_db.execute(
            select(SavedReport).where(SavedReport.id == uuid.UUID(saved_report_id))
        )
        saved = result.scalar_one_or_none()
        if not saved:
            logger.error(f"Scheduled report {report_id}: saved report {saved_report_id} not found")
            return
        config_dict = saved.config
        report_name = saved.name

    # Build and run the query
    config = ReportConfig(**config_dict)
    sql_str, params, select_parts = _build_report_query(config)

    async with OdooSessionLocal() as odoo_db:
        result = await odoo_db.execute(text(sql_str), params)
        rows = [dict(r) for r in result.mappings().all()]

    # Derive column keys/headers
    headers, keys = [], []
    for sp in select_parts:
        if " as " in sp:
            key = sp.split(" as ")[-1].strip()
        elif "." in sp:
            key = sp.split(".")[-1].strip()
        else:
            key = sp.strip()
        keys.append(key)
        headers.append(key.replace("_", " ").title())

    table_label = AVAILABLE_TABLES.get(config.table, {}).get("label", config.table)

    # Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = table_label[:31]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    attachments = [(
        f"report_builder_{config.table}.xlsx",
        excel_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )]

    subject = f"RT Reporting - {report_name} ({date_from} to {date_to})"
    html_body = f"""<html><body>
    <h2>{report_name}</h2>
    <p>Your scheduled report has been generated with <strong>{len(rows):,}</strong> records.</p>
    <p>Table: {table_label} &bull; Period: {date_from} to {date_to}</p>
    <p>See the attached Excel file for the full data.</p>
    </body></html>"""

    try:
        await send_email(recipients, subject, html_body, attachments)
        logger.info(f"Sent report_builder report '{report_name}' to {recipients}")
    except Exception as e:
        logger.error(f"Failed to send report_builder report: {e}")


async def _generate_excel_attachment(db, report_type: str, date_from: str, date_to: str) -> bytes | None:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.title()

        if report_type == "sales":
            result = await sales_service.get_sales_orders(db, date.fromisoformat(date_from), date.fromisoformat(date_to), limit=500)
            headers = ["Order #", "Customer", "Date", "Total", "Status"]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)
            for i, o in enumerate(result["orders"], 2):
                ws.cell(row=i, column=1, value=o.get("name"))
                ws.cell(row=i, column=2, value=o.get("customer_name"))
                ws.cell(row=i, column=3, value=o.get("date_order"))
                ws.cell(row=i, column=4, value=o.get("amount_total"))
                ws.cell(row=i, column=5, value=o.get("state"))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        return None
